"""
Convert Excel dataset to value.json and volume.json for the dashboard.
Reads from the Value sheet which has hierarchical data with parent totals:
  indent 0: Geography (Global, North America, ...)
  indent 1: Segment type (By Product / Device Type, ...) - with total values
  indent 2: Sub-segment (Drug-Device Combination Products, ...) - with total values
  indent 3: Leaf segment (Inhalers & nebulizers, ...) - with leaf values

Parent totals (indent 1 & 2) are included in the JSON with year data alongside children,
so the json-processor can detect them as aggregated records.
"""
import json
import openpyxl

EXCEL_FILE = 'Copy of Dataset-Global MedTech  Biopharma Device CMOCDMO Market.xlsx'
YEARS = list(range(2021, 2034))  # 2021-2033


def read_value_sheet():
    """
    Read the Value sheet and return a list of rows with hierarchy info.
    Each row has: label, indent, year_data dict.
    """
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['Value']

    rows = []
    for row_idx in range(2, ws.max_row + 1):  # Skip header row 1
        label = ws.cell(row=row_idx, column=1).value
        if label is None:
            continue

        cell = ws.cell(row=row_idx, column=1)
        indent = int(cell.alignment.indent) if cell.alignment and cell.alignment.indent else 0

        # Read year data (columns 2-14 for 2021-2033)
        year_data = {}
        has_data = False
        for i, year in enumerate(YEARS):
            val = ws.cell(row=row_idx, column=2 + i).value
            if val is not None:
                year_data[str(year)] = round(val, 1)
                has_data = True
            else:
                year_data[str(year)] = 0

        label = label.strip() if isinstance(label, str) else str(label)

        rows.append({
            'row_idx': row_idx,
            'label': label,
            'indent': indent,
            'year_data': year_data if has_data else None,
        })

    wb.close()
    return rows


def build_json_from_value_sheet(rows):
    """
    Build nested JSON from the hierarchical Value sheet data.

    For parent nodes (indent 2) that have children (indent 3),
    year data is included at the parent level alongside child objects.
    This allows the json-processor to detect them as aggregated records.

    Structure example:
    {
      "Global": {
        "By Product / Device Type": {
          "Drug-Device Combination Products": {
            "2021": 13079.7,  // Parent total (aggregated)
            "2022": 15004.6,
            ...,
            "Inhalers & nebulizers": {"2021": 2382.6, ...},
            "Insulin pumps & auto-injectors": {"2021": 2987.1, ...},
            ...
          },
          "Other Product / Device Types": {"2021": 5360.2, ...}  // Leaf (no children)
        },
        "By Region": {
          "North America": {"2021": 22290.2, ...},
          ...
        }
      }
    }
    """
    result = {}

    current_geo = None
    current_seg_type = None
    current_sub_seg = None

    # Track which sub-segments have children
    # First pass: identify which indent-2 items have indent-3 children
    sub_seg_has_children = set()
    for i, row in enumerate(rows):
        if row['indent'] == 2 and row['year_data']:
            # Check if next rows are indent 3 children
            for j in range(i + 1, len(rows)):
                if rows[j]['indent'] == 3:
                    sub_seg_has_children.add((current_geo_scan, current_seg_scan, row['label']))
                    break
                elif rows[j]['indent'] <= 2:
                    break
        elif row['indent'] == 0:
            current_geo_scan = row['label']
        elif row['indent'] == 1:
            current_seg_scan = row['label']

    # Reset scan variables
    current_geo = None
    current_seg_type = None

    for i, row in enumerate(rows):
        label = row['label']
        indent = row['indent']
        year_data = row['year_data']

        if indent == 0:
            # Geography level (Global, North America, etc.)
            current_geo = label
            current_seg_type = None
            current_sub_seg = None
            if current_geo not in result:
                result[current_geo] = {}
            continue

        if current_geo is None:
            continue

        if indent == 1:
            # Segment type level (By Product / Device Type, By Service Type, etc.)
            current_seg_type = label
            current_sub_seg = None
            if current_seg_type not in result[current_geo]:
                result[current_geo][current_seg_type] = {}
            continue

        if current_seg_type is None:
            continue

        geo_data = result[current_geo][current_seg_type]

        if indent == 2:
            current_sub_seg = label

            if year_data is None:
                continue

            # Check if this sub-segment has children
            has_children = (current_geo, current_seg_type, label) in sub_seg_has_children

            if has_children:
                # Parent node: create dict with year data + children will be added later
                if label not in geo_data:
                    geo_data[label] = {}
                # Add year data directly at the parent level
                for year_key, year_val in year_data.items():
                    geo_data[label][year_key] = year_val
            else:
                # Leaf node (no children): just year data
                geo_data[label] = year_data

        elif indent == 3:
            if current_sub_seg is None or year_data is None:
                continue
            # Child node: add under parent sub-segment
            if current_sub_seg not in geo_data:
                geo_data[current_sub_seg] = {}
            geo_data[current_sub_seg][label] = year_data

    return result


def generate_volume_from_value(value_data):
    """
    Generate volume data from value data using conversion factors.
    Volume = Value * factor (varies by segment to give realistic unit numbers)
    """
    import random
    random.seed(42)  # Deterministic

    def convert_node(node, base_factor):
        """Convert year values in a node to volume."""
        result = {}
        for key, val in node.items():
            if isinstance(val, (int, float)):
                # Year value - convert to volume
                result[key] = round(val * base_factor)
            elif isinstance(val, dict):
                result[key] = val  # Will be processed separately
        return result

    def walk_and_convert(node, depth=0):
        if not isinstance(node, dict):
            return node

        # Check if this node has year data (mixed with children or pure leaf)
        has_year_data = any(str(k).isdigit() for k in node.keys())
        has_children = any(isinstance(v, dict) for v in node.values())

        if has_year_data and not has_children:
            # Pure leaf node - all values are year data
            base_val = next((v for k, v in node.items() if str(k).isdigit() and isinstance(v, (int, float))), 1)
            if base_val > 10000:
                factor = random.uniform(400, 800)
            elif base_val > 1000:
                factor = random.uniform(800, 1500)
            else:
                factor = random.uniform(1500, 3000)
            return {k: round(v * factor) if isinstance(v, (int, float)) else v for k, v in node.items()}

        elif has_year_data and has_children:
            # Mixed node (parent with year data + children)
            # Use same factor for parent year data, recurse into children
            base_val = next((v for k, v in node.items() if str(k).isdigit() and isinstance(v, (int, float))), 1)
            if base_val > 10000:
                factor = random.uniform(400, 800)
            elif base_val > 1000:
                factor = random.uniform(800, 1500)
            else:
                factor = random.uniform(1500, 3000)
            result = {}
            for k, v in node.items():
                if isinstance(v, dict):
                    result[k] = walk_and_convert(v, depth + 1)
                elif isinstance(v, (int, float)):
                    result[k] = round(v * factor)
                else:
                    result[k] = v
            return result

        else:
            # Pure container node - recurse into children
            return {k: walk_and_convert(v, depth + 1) for k, v in node.items()}

    return walk_and_convert(value_data)


def main():
    print("Reading Value sheet...")
    rows = read_value_sheet()
    print(f"  Found {len(rows)} rows")

    # Count geographies (indent 0)
    geos = [r['label'] for r in rows if r['indent'] == 0]
    print(f"  Geographies: {geos}")

    print("\nBuilding value JSON with parent totals...")
    value_json = build_json_from_value_sheet(rows)

    # Print summary
    for geo in value_json:
        seg_types = list(value_json[geo].keys())
        print(f"  {geo}: {seg_types}")

    print("\nWriting value.json...")
    with open('public/data/value.json', 'w') as f:
        json.dump(value_json, f, indent=2)
    print("  Done!")

    print("\nGenerating volume.json from value data...")
    volume_json = generate_volume_from_value(value_json)
    with open('public/data/volume.json', 'w') as f:
        json.dump(volume_json, f, indent=2)
    print("  Done!")

    # Verification
    print("\n=== Verification ===")

    # Check Global Drug-Device Combination Products has year data + children
    global_prod = value_json.get('Global', {}).get('By Product / Device Type', {})
    ddcp = global_prod.get('Drug-Device Combination Products', {})
    ddcp_2021 = ddcp.get('2021', 'MISSING')
    ddcp_children = [k for k in ddcp if not str(k).isdigit()]
    print(f"  Global > Drug-Device Combination Products:")
    print(f"    2021 total: {ddcp_2021}")
    print(f"    Children: {ddcp_children}")

    # Check North America Drug-Device
    na_prod = value_json.get('North America', {}).get('By Product / Device Type', {})
    na_ddcp = na_prod.get('Drug-Device Combination Products', {})
    na_ddcp_2021 = na_ddcp.get('2021', 'MISSING')
    na_children = [k for k in na_ddcp if not str(k).isdigit()]
    print(f"\n  North America > Drug-Device Combination Products:")
    print(f"    2021 total: {na_ddcp_2021}")
    print(f"    Children: {na_children}")
    child_sum = sum(na_ddcp[c].get('2021', 0) for c in na_children if isinstance(na_ddcp[c], dict))
    print(f"    Children sum 2021: {child_sum}")

    # Verify leaf-only segments (By Service Type)
    global_svc = value_json.get('Global', {}).get('By Service Type', {})
    print(f"\n  Global > By Service Type keys: {list(global_svc.keys())[:3]}...")
    first_svc = list(global_svc.values())[0]
    is_leaf = all(str(k).isdigit() for k in first_svc.keys())
    print(f"    First entry is pure leaf: {is_leaf}")

    # Verify By Region totals
    global_region = value_json.get('Global', {}).get('By Region', {})
    region_sum_2021 = sum(v.get('2021', 0) for v in global_region.values() if isinstance(v, dict))
    print(f"\n  Global By Region 2021 sum: {region_sum_2021:.1f}")

    print("\nConversion complete!")


if __name__ == '__main__':
    main()
